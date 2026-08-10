"""
Build a submission-ready, properly formatted Excel workbook
from companies_dataset.csv.

Produces: IT_Prospecting_Pipeline_Pakistan.xlsx
Sheets:
    1. Leads         - main dataset (styled, filtered, frozen header)
    2. Summary       - counts by vertical + status
    3. Candidate     - candidate profile + resume reference
    4. Instructions  - how the pipeline is used
"""

import csv
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from email_templates import EMAIL_TEMPLATES


BASE_DIR = Path(__file__).parent
CSV_FILE = BASE_DIR / "companies_dataset.csv"
XLSX_FILE = BASE_DIR / "IT_Prospecting_Pipeline_Pakistan.xlsx"

DEFAULT_RESUME = "Usairam Saeed.pdf"
AI_RESUME = "Usairam_Saeed_EuroPass_CV.pdf"
CANDIDATE_EMAIL = "saeed.usairam@gmail.com"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F4E78")
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="404040")
LABEL_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E78")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

VERTICAL_COLORS = {
    "E-Commerce & Retail":    "FFF2CC",
    "Education & Training":   "DEEBF7",
    "Marketing & Advertising":"FCE4D6",
    "AI & Data Science":      "E2EFDA",
    "Travel & Hospitality":   "EAD1DC",
}

STATUS_COLORS = {
    "Pending":     "FFF2CC",
    "Applied":     "DEEBF7",
    "Interview":   "E2EFDA",
    "Rejected":    "F8CBAD",
    "Offer":       "C6E0B4",
    "Failed":      "D9D9D9",
}

# Apply method palette: ATS is the most reliable channel, Email the most common
# for Pakistan's SME IT sector, Portal requires a form fill, Unknown needs review.
APPLY_METHOD_COLORS = {
    "ATS":      "C6E0B4",
    "Portal":   "DEEBF7",
    "Email":    "FFF2CC",
    "Unknown":  "F8CBAD",
}


def apply_method_color(value):
    """APPLY_METHOD values look like 'ATS (Greenhouse)' - key on the prefix."""
    if not value:
        return None
    for key, color in APPLY_METHOD_COLORS.items():
        if value.startswith(key):
            return color
    return None


def load_rows():
    with open(CSV_FILE, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)
    return rows[0], rows[1:]


def style_leads_sheet(ws, headers, data):
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value="IT Prospecting Pipeline - Pakistan").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(
        row=2, column=1,
        value=(
            "Candidate: Usairam Saeed  |  Full Stack Product Engineer (React, Node.js, MongoDB)  "
            f"|  Email: {CANDIDATE_EMAIL}  |  Country focus: Pakistan"
        ),
    ).font = SUBTITLE_FONT
    ws.row_dimensions[2].height = 20

    # Empty spacer row
    ws.row_dimensions[3].height = 6

    header_row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[header_row].height = 32

    status_idx = headers.index("Application Status") if "Application Status" in headers else -1
    # Data rows
    for row_idx, row in enumerate(data, start=header_row + 1):
        vertical = row[0] if row else ""
        status = row[status_idx] if status_idx >= 0 and len(row) > status_idx else ""
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = LEFT
            cell.border = BORDER
            cell.font = Font(name="Calibri", size=10)
            header_name = headers[col_idx - 1]
            if header_name == "Vertical" and vertical in VERTICAL_COLORS:
                cell.fill = PatternFill("solid", fgColor=VERTICAL_COLORS[vertical])
                cell.font = Font(name="Calibri", size=10, bold=True)
            if header_name == "Application Status" and status in STATUS_COLORS:
                cell.fill = PatternFill("solid", fgColor=STATUS_COLORS[status])
                cell.alignment = CENTER
                cell.font = Font(name="Calibri", size=10, bold=True)
            if header_name == "Apply Method":
                color = apply_method_color(value)
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)
                    cell.alignment = CENTER
                    cell.font = Font(name="Calibri", size=10, bold=True)
            if header_name in (
                "Website", "Careers Page", "Contact Page", "Source URL",
                "Rozee.pk Search", "LinkedIn Jobs",
            ) and value:
                cell.hyperlink = value
                cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
            if header_name in ("Apply Email", "Info Email", "HR Email", "Candidate Email") and value:
                cell.hyperlink = f"mailto:{value}"
                cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
            if header_name == "Resume to Send" and value:
                cell.hyperlink = value
                cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

    # Column widths
    widths = {
        "Vertical": 24,
        "Organization Name": 24,
        "Website": 30,
        "Careers Page": 34,
        "ATS Platform": 16,
        "Apply Method": 22,
        "Contact Page": 30,
        "Apply Email": 28,
        "Info Email": 26,
        "HR Email": 26,
        "Rozee.pk Search": 34,
        "LinkedIn Jobs": 34,
        "Source URL": 32,
        "Country": 12,
        "Notes": 45,
        "Custom Requirement": 40,
        "Resume to Send": 34,
        "Candidate Email": 26,
        "Application Status": 18,
    }
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(header, 20)

    # Freeze header
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Auto-filter
    last_col_letter = get_column_letter(len(headers))
    last_row = header_row + len(data)
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_row}"

    # Proper Excel Table for nice formatting
    table_ref = f"A{header_row}:{last_col_letter}{last_row}"
    table = Table(displayName="LeadsTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    # NOTE: Using either Table OR manual styling. Tables in openpyxl override cell fills,
    # so we skip adding the table if we want to keep per-row colored fills.
    # ws.add_table(table)  # intentionally disabled to preserve vertical color coding


def build_summary_sheet(ws, headers, data):
    ws.merge_cells("A1:D1")
    ws["A1"] = "Summary - IT Prospecting Pipeline"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 28

    # Totals by vertical
    ws["A3"] = "Companies by Vertical"
    ws["A3"].font = LABEL_FONT

    ws["A4"], ws["B4"] = "Vertical", "Count"
    for c in ("A4", "B4"):
        ws[c].fill = HEADER_FILL
        ws[c].font = HEADER_FONT
        ws[c].alignment = CENTER
        ws[c].border = BORDER

    vertical_counts = Counter(row[0] for row in data)
    row_idx = 5
    for vertical, count in vertical_counts.most_common():
        ws.cell(row=row_idx, column=1, value=vertical).border = BORDER
        ws.cell(row=row_idx, column=1).fill = PatternFill("solid", fgColor=VERTICAL_COLORS.get(vertical, "FFFFFF"))
        ws.cell(row=row_idx, column=1).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=row_idx, column=2, value=count).alignment = CENTER
        ws.cell(row=row_idx, column=2).border = BORDER
        row_idx += 1

    total_row = row_idx
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=sum(vertical_counts.values())).font = Font(bold=True)
    for col in (1, 2):
        ws.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor="D9E1F2")
        ws.cell(row=total_row, column=col).border = BORDER
        ws.cell(row=total_row, column=col).alignment = CENTER

    # Totals by status
    status_start = total_row + 3
    ws.cell(row=status_start, column=1, value="Companies by Application Status").font = LABEL_FONT

    ws.cell(row=status_start + 1, column=1, value="Status")
    ws.cell(row=status_start + 1, column=2, value="Count")
    for col in (1, 2):
        c = ws.cell(row=status_start + 1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER

    status_idx = headers.index("Application Status") if "Application Status" in headers else -1
    status_counts = Counter(row[status_idx] for row in data if status_idx >= 0 and len(row) > status_idx)
    rr = status_start + 2
    for status, count in status_counts.most_common():
        ws.cell(row=rr, column=1, value=status).border = BORDER
        ws.cell(row=rr, column=1).fill = PatternFill("solid", fgColor=STATUS_COLORS.get(status, "FFFFFF"))
        ws.cell(row=rr, column=1).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=rr, column=1).alignment = CENTER
        ws.cell(row=rr, column=2, value=count).alignment = CENTER
        ws.cell(row=rr, column=2).border = BORDER
        rr += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14


def build_candidate_sheet(ws):
    ws.merge_cells("A1:B1")
    ws["A1"] = "Candidate Profile"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 28

    rows = [
        ("Full Name", "Usairam Saeed", None),
        ("Role", "Full Stack Product Engineer", None),
        ("Stack", "React, Node.js, MongoDB", None),
        ("Experience", "SaaS platforms with 500 - 2000+ users", None),
        ("Specialisations", "Payments, real-time systems, scheduling", None),
        ("AI Work", "AI-integrated HR SaaS (high accuracy, large-scale usage)", None),
        ("UI/UX", "Strong UI/UX background with measurable improvements", None),
        ("Country Focus", "Pakistan", None),
        ("Target Verticals", "E-Commerce & Retail, Education & Training, Marketing & Advertising, AI & Data Science, Travel & Hospitality", None),
        ("Candidate Email", CANDIDATE_EMAIL, f"mailto:{CANDIDATE_EMAIL}"),
        ("Default Resume", DEFAULT_RESUME, DEFAULT_RESUME),
        ("Default Resume - Used For", "E-Commerce & Retail, Education & Training, Marketing & Advertising, Travel & Hospitality", None),
        ("AI Resume (Europass)", AI_RESUME, AI_RESUME),
        ("AI Resume - Used For", "AI & Data Science", None),
    ]

    for idx, (label, value, link) in enumerate(rows, start=3):
        lc = ws.cell(row=idx, column=1, value=label)
        vc = ws.cell(row=idx, column=2, value=value)
        lc.font = LABEL_FONT
        lc.fill = PatternFill("solid", fgColor="D9E1F2")
        lc.alignment = LEFT
        lc.border = BORDER
        vc.alignment = LEFT
        vc.border = BORDER
        vc.font = Font(name="Calibri", size=11)
        if link:
            vc.hyperlink = link
            vc.font = Font(name="Calibri", size=11, color="0563C1", underline="single")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90
    for r in range(3, 3 + len(rows)):
        ws.row_dimensions[r].height = 22


def build_instructions_sheet(ws):
    ws.merge_cells("A1:B1")
    ws["A1"] = "How to Use This Workbook"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 28

    lines = [
        ("1. Leads sheet", "Main prospecting dataset. Use filters on each column to narrow by Vertical, Country or Application Status."),
        ("2. Summary sheet", "Auto-counts by Vertical and Application Status for a quick snapshot."),
        ("3. Candidate sheet", "Candidate profile, email (saeed.usairam@gmail.com) and links to both resume files."),
        ("4. Cold Email Templates", "10 copy-paste outreach templates per vertical + follow-up + referral request. Replace the {placeholders}."),
        ("5. Resume to Send", "AI & Data Science rows use Usairam_Saeed_EuroPass_CV.pdf. All other verticals use Usairam Saeed.pdf."),
        ("6. Candidate Email", "saeed.usairam@gmail.com is included on every row as the reply-to address."),
        ("7. Application Status", "Pending, Applied, Interview, Rejected, Offer, Failed (bad/undeliverable email — set back to Pending after fixing). Rows are color-coded."),
        ("8. Pipeline script", "prospecting_pipeline.py enriches careers/contact pages, ATS, Rozee.pk and LinkedIn URLs, then pushes to Google Sheets."),
        ("9. Auto-sender", "send_applications.py reads this CSV, picks the right template + resume per vertical, sends via Gmail SMTP, throttles, and updates status to Applied."),
        ("10. Safe test", "Use 'python send_applications.py --dry-run' to preview and '--to-self' to route every email to your own inbox first."),
    ]
    for idx, (label, value) in enumerate(lines, start=3):
        lc = ws.cell(row=idx, column=1, value=label)
        vc = ws.cell(row=idx, column=2, value=value)
        lc.font = LABEL_FONT
        lc.fill = PatternFill("solid", fgColor="D9E1F2")
        lc.alignment = LEFT
        lc.border = BORDER
        vc.alignment = LEFT
        vc.border = BORDER
        ws.row_dimensions[idx].height = 34

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 110


def build_templates_sheet(ws):
    """Cold-email templates the candidate can copy-paste when reaching out."""
    ws.merge_cells("A1:D1")
    ws["A1"] = "Cold Email Templates"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    ws["A2"] = (
        "Replace placeholders like {Company}, {Vertical}, {HiringManager}, "
        "{OneLineAboutCompany}, {ContactName} before sending. Keep it short and specific."
    )
    ws["A2"].font = SUBTITLE_FONT
    ws.row_dimensions[2].height = 22

    header_row = 4
    headers = ["ID", "Template Name", "Best Used For", "Subject", "Body"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[header_row].height = 30

    row_idx = header_row + 1
    for template in EMAIL_TEMPLATES:
        ws.cell(row=row_idx, column=1, value=template["id"])
        ws.cell(row=row_idx, column=2, value=template["name"])
        ws.cell(row=row_idx, column=3, value=template["use_for"])
        ws.cell(row=row_idx, column=4, value=template["subject"])
        ws.cell(row=row_idx, column=5, value=template["body"])

        for col_idx in range(1, 6):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = BORDER
            cell.font = Font(name="Calibri", size=10)
            if col_idx == 5:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.font = Font(name="Calibri", size=10, bold=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        # Approximate row height so each template is readable
        line_count = template["body"].count("\n") + 2
        ws.row_dimensions[row_idx].height = max(90, line_count * 14)
        row_idx += 1

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 48
    ws.column_dimensions["E"].width = 95

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def main():
    headers, data = load_rows()

    wb = Workbook()

    ws_leads = wb.active
    ws_leads.title = "Leads"
    style_leads_sheet(ws_leads, headers, data)

    ws_summary = wb.create_sheet("Summary")
    build_summary_sheet(ws_summary, headers, data)

    ws_candidate = wb.create_sheet("Candidate")
    build_candidate_sheet(ws_candidate)

    ws_templates = wb.create_sheet("Cold Email Templates")
    build_templates_sheet(ws_templates)

    ws_instructions = wb.create_sheet("Instructions")
    build_instructions_sheet(ws_instructions)

    wb.active = 0
    try:
        wb.save(XLSX_FILE)
        print(f"[OK] Workbook saved: {XLSX_FILE}")
    except PermissionError:
        fallback = XLSX_FILE.with_name(XLSX_FILE.stem + "_new.xlsx")
        wb.save(fallback)
        print(f"[WARN] {XLSX_FILE.name} is open in Excel - saved to {fallback.name} instead.")
        print("       Close Excel and re-run to overwrite the original file.")


if __name__ == "__main__":
    main()
